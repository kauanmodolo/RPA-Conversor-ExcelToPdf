import openpyxl
from pathlib import Path
from models.nota_fiscal_model import NotaFiscalModel

class Excel_Manager:
    def __init__(self, caminho_excel: str):
        
        """
        Inicializa o gerenciador de Excel.

        Args:
            caminho_excel (str): Caminho para o arquivo Excel.
        """

        self.caminho_excel = Path(caminho_excel)
        self.wb = None
        self.ws = None
        self.colunas_validas = None

    def carregar_excel(self) -> None:
        
        """
        Carrega o arquivo Excel e armazena a planilha ativa e as colunas válidas.

        Raises:
            FileNotFoundError: Se o arquivo Excel não for encontrado.
        """

        print(f"\n📂 Carregando arquivo Excel: {self.caminho_excel.name}...")
        if not self.caminho_excel.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.caminho_excel}")
        
        # Abre o arquivo Excel e seleciona a planilha ativa
        self.wb = openpyxl.load_workbook(self.caminho_excel, data_only=True)
        self.ws = self.wb.active
        
        # Lê o cabeçalho (primeira linha) e armazena os nomes das colunas
        self.colunas_validas = [cell.value.strip() if isinstance(cell.value, str) else str(cell.value) for cell in next(self.ws.iter_rows(min_row=1, max_row=1))]
        print(f"✅ Excel carregado. Colunas disponíveis: {', '.join(self.colunas_validas)}")

    def buscar_contrato(self, numero_contrato: int, coluna: str = "CONTRATO") -> list:
        
        """
        Busca registros no Excel pelo número do contrato em uma coluna específica e retorna uma lista de NotaFiscalModel.

        Args:
            numero_contrato (int): Número do contrato a ser buscado.
            coluna (str, opcional): Nome da coluna onde buscar. Padrão: "CONTRATO".

        Returns:
            list: Lista de instâncias NotaFiscalModel encontradas.

        Raises:
            RuntimeError: Se a planilha não estiver carregada.
            ValueError: Se a coluna não existir ou nenhum registro for encontrado.
        """

        print(f"\n🔍 Buscando contrato {numero_contrato} na coluna '{coluna}'...")
        if self.ws is None:
            raise RuntimeError("Planilha não carregada. Execute carregar_excel() primeiro")
            
        if coluna not in self.colunas_validas:
            raise ValueError(f"Coluna '{coluna}' não existe. Colunas válidas: {self.colunas_validas}")
        
        # Descobre o índice da coluna onde será feita a busca
        col_idx = self.colunas_validas.index(coluna)
        resultados = []
        
        # Percorre todas as linhas da planilha (exceto o cabeçalho)
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            valor_celula = row[col_idx]
            try:
                # Verifica se o valor da célula corresponde ao número do contrato procurado
                if valor_celula is not None and str(valor_celula).isdigit() and int(valor_celula) == numero_contrato:
                    
                    # Cria um dicionário associando cada coluna ao seu valor na linha
                    row_data = {self.colunas_validas[i]: row[i] for i in range(len(self.colunas_validas))}
                    
                    # Instancia o modelo NotaFiscalModel com os dados da linha
                    nota_fiscal = NotaFiscalModel(
                        CONTRATO=int(row_data.get("CONTRATO", 0)),
                        CODIGO_PARCEIRO=int(row_data.get("CODIGO PARCEIRO", 0)),
                        TIPO_DE_NEGOCIACAO=int(row_data.get("TIPO DE NEGOCIACAO", 0)),
                        TIPO_OPERACAO=int(row_data.get("TIPO OPERACAO", 0)),
                        CENTRO_RESULTADO=int(row_data.get("CENTRO RESULTADO", 0)),
                        PROJETO=int(row_data.get("PROJETO", 0)),
                        NATUREZA=int(row_data.get("NATUREZA", 0)),
                        CIDADE=int(row_data.get("CIDADE", 0)),
                        CIDADE_SERVICO=int(row_data.get("CIDADE SERVICO", 0)),
                        PRODUTO=int(row_data.get("PRODUTO", 0)),
                    )
                    resultados.append(nota_fiscal)
            except (ValueError, TypeError) as e:
                print(f"Erro ao processar linha: {e}")
                continue
        
        # Se nenhum registro for encontrado, lança exceção
        if not resultados:
            raise ValueError(f"Nenhum registro encontrado para o contrato {numero_contrato}")
            
        print(f"✅ Contrato encontrado com {len(resultados)} registro(s)")
        return resultados