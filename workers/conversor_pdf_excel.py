import pdfplumber
import openpyxl
from typing import Dict
# import tempfile  # Descomente para usar arquivo temporário

class Conversor_Pdf_Excel:
    def __init__(self, pdf_path, excel_path):
        
        """
        Inicializa o extrator de tabelas de PDF para Excel.

        Args:
            pdf_path (str): Caminho do arquivo PDF de entrada.
            excel_path (str): Caminho do arquivo Excel de saída.
        """

        self.pdf_path = pdf_path
        self.excel_path = excel_path
        self.all_tables = []

        # Para gerar arquivo Excel temporário, descomente abaixo:
        # temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        # self.excel_path = temp_file.name

    def extrair_tabelas(self):
        
        """
        Extrai todas as tabelas de todas as páginas do PDF e armazena em self.all_tables.

        Returns:
            None
        """
        
        print(f"🔍 Abrindo PDF: {self.pdf_path}")
        with pdfplumber.open(self.pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                print(f"  ➡️  Extraindo tabelas da página {page_num}...")
                tables = page.extract_tables()
                for t_idx, table in enumerate(tables, 1):
                    print(f"    ✔️  Tabela {t_idx} extraída da página {page_num} (linhas: {len(table)})")
                    self.all_tables.append(table)
        print(f"✅ Total de tabelas extraídas: {len(self.all_tables)}")

    def salvar_para_excel(self):
        
        """
        Salva todas as tabelas extraídas em um arquivo Excel, cada tabela em uma aba diferente.

        Returns:
            None
        
        """

        print(f"💾 Salvando tabelas no arquivo Excel: {self.excel_path}")
        wb = openpyxl.Workbook()
        
        # Remove a planilha padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Adiciona cada tabela em uma aba diferente
        for i, table in enumerate(self.all_tables):
            sheet = wb.create_sheet(title=f'Tabela_{i+1}')
            for row in table:
                sheet.append(row)
            print(f"  📝 Tabela_{i+1} salva com {len(table)} linhas.")
        wb.save(self.excel_path)
        print("✅ Arquivo Excel salvo com sucesso.")

    def executar(self):
        
        """
        Executa o processo completo de extração das tabelas do PDF e salva no Excel.

        Returns:
            None
        
        """

        self.extrair_tabelas()
        self.salvar_para_excel()

    def encontrar_dados_no_excel(self) -> Dict:
        
        """
        Busca adaptativa no Excel gerado, percorrendo todas as abas e identificando
        as colunas de 'contrato', 'conceito' e 'valor' independentemente da posição.

        Returns:
            dict: Dicionário com os dados encontrados: 'contrato_principal', 'valor_total' e 'conceito'.
        """

        resultados = {
            "arquivo": self.excel_path,
            "contrato_principal": None,
            "valor_total": None,
            "conceito": None,
        }

        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            for sheet_name in wb.sheetnames:
                print(f"\n🔎 Lendo aba: '{sheet_name}'")
                sheet = wb[sheet_name]
                
                # Define o que é o header e pega o cabeçalho da primeira linha
                header = [str(cell.value).strip().lower() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                print(f"   ➡️ Cabeçalho encontrado: {header}")
                
                # Identifica os índices de interesse
                idx_contrato = next((i for i, v in enumerate(header) if "contrato" in v), None)
                idx_conceito = next((i for i, v in enumerate(header) if "conceito" in v), None)
                idx_valor = next((i for i, v in enumerate(header) if "valor" in v), None)
                print(f"   ➡️ Índices encontrados - Contrato: {idx_contrato}, Conceito: {idx_conceito}, Valor: {idx_valor}")
                
                # Percorre as linhas buscando os valores
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    print(f"      ➡️ Lendo linha {row_num}': {row}")
                    
                    # Contrato
                    if idx_contrato is not None and not resultados["contrato_principal"]:
                        contrato = row[idx_contrato]
                        if contrato and str(contrato).isdigit():
                            resultados["contrato_principal"] = int(contrato)
                            print(f"         ✔️ Contrato encontrado na linha {row_num}: {resultados['contrato_principal']}")
                    
                    # Conceito
                    if idx_conceito is not None and not resultados["conceito"]:
                        conceito = row[idx_conceito]
                        if conceito:
                            resultados["conceito"] = str(conceito).strip()
                            print(f"         ✔️ Conceito encontrado na linha {row_num}: {resultados['conceito']}")
                    
                    # Valor
                    if idx_valor is not None and not resultados["valor_total"]:
                        valor = row[idx_valor]
                        if valor:
                            try:
                                valor_str = str(valor).replace('"', '').replace('.', '').replace(',', '.')
                                resultados["valor_total"] = float(valor_str)
                                print(f"         ✔️ Valor encontrado na linha {row_num}: {resultados['valor_total']}")
                            except Exception:
                                resultados["valor_total"] = valor
                                print(f"         ✔️ Valor encontrado na linha {row_num} (não convertido): {resultados['valor_total']}")
                    
                    # Se já encontrou tudo, pode parar
                    if resultados["contrato_principal"] and resultados["conceito"] and resultados["valor_total"]:
                        print("      ✅ Todos os dados encontrados nesta aba.")
                        break
                
                # Se já encontrou tudo, pode parar de buscar em outras abas
                if resultados["contrato_principal"] and resultados["conceito"] and resultados["valor_total"]:
                    print("✅ Dados completos encontrados. Parando busca nas abas.")
                    break
        
        except Exception as e:
            print(f"Erro ao ler Excel: {str(e)}")

        print(f"\n📋 Resultado final extraído do Excel: {resultados}")
        return resultados